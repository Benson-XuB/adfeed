"""生成全部 1688 卖家常见 Excel 格式（仿真数据，用于测试 _ingest_universal.py）

覆盖 8 种真实来源 + 6 个品类。每份 15-25 行，含常见脏数据。

用法：
    python3 _gen_all_excels.py                     # 输出到 ./mock_excels/
    python3 _gen_all_excels.py --dir /tmp/mocks
"""
import os, sys, csv, random, openpyxl
from openpyxl.styles import Font, Alignment
from pathlib import Path

OUT_DIR = Path(os.path.dirname(os.path.abspath(__file__))) / "mock_excels"

# ── 真实产品数据（6 大类，含材质/颜色/描述） ──
_PRODUCTS = {
    "美容个护": [
        ("玻尿酸补水面膜 28ml*10片保湿", 39.9, 500, "玻尿酸 蚕丝膜布", "透明", "补水保湿 淡化细纹"),
        ("烟酰胺亮肤美白精华液 30ml", 59.0, 320, "烟酰胺 透明质酸", "乳白色", "美白淡斑 提亮肤色"),
        ("氨基酸温和洁面泡沫 150ml", 45.0, 600, "氨基酸 甘油", "白色泡沫", "敏感肌可用 深层清洁"),
        ("视黄醇抗皱紧致眼霜 20g", 78.0, 250, "视黄醇 胜肽", "淡黄色", "淡化眼纹 紧致眼周"),
        ("积雪草舒缓修护面霜 50g", 89.0, 180, "积雪草 神经酰胺", "绿色", "舒缓泛红 修复屏障"),
        ("胶原蛋白睡眠唇膜 15g", 29.9, 800, "胶原蛋白 蜂蜜", "粉色", "夜间修护 淡化唇纹"),
        ("竹炭深层清洁泥膜 100g", 49.0, 400, "竹炭粉 高岭土", "黑色", "吸附黑头 控油清洁"),
        ("防晒喷雾SPF50+ PA++++ 150ml", 69.0, 300, "物理防晒 氧化锌", "白色雾状", "高倍防晒 清爽不腻"),
        ("玫瑰花瓣沐浴露 500ml", 49.0, 420, "玫瑰花瓣 甘油", "粉色", "持久留香 温和清洁"),
        ("茶树精油祛痘精华 10ml", 39.0, 500, "茶树精油 水杨酸", "透明", "快速消痘 抑菌消炎"),
        ("维生素C亮白精华液 30ml", 55.0, 280, "维C 阿魏酸", "淡黄色", "抗氧化亮白 均匀肤色"),
        ("玻尿酸保湿喷雾 100ml", 42.0, 600, "玻尿酸 温泉水", "透明", "随时补水 舒缓镇定 定妆"),
        ("胶原蛋白弹力颈霜 50g", 68.0, 200, "胶原蛋白 弹力蛋白", "白色", "提拉紧致 淡化颈纹"),
    ],
    "珠宝饰品": [
        ("S925纯银锁骨链女轻奢小众项链", 29.9, 200, "S925银 锆石", "银色", "不掉色 不过敏 精致日常"),
        ("18K镀金极细手链女ins风轻奢", 39.0, 150, "钛钢 镀18K金", "金色", "防水不掉色 简约百搭"),
        ("天然淡水珍珠耳钉女气质耳饰", 59.0, 120, "天然珍珠 S925银针", "白色", "正圆强光 不过敏 优雅"),
        ("复古锆石戒指女开口可调食指戒", 19.9, 400, "合金 锆石", "银色 金色", "开口设计 不挑手 闪耀"),
        ("韩版蝴蝶结发夹少女甜美刘海夹", 9.9, 800, "合金 亚克力", "粉色 黑色", "一字夹百搭 不滑落"),
        ("钛钢项链男嘻哈古巴链不掉色", 49.0, 250, "316L钛钢", "银色", "不掉色 不过敏 硬汉风"),
        ("天然水晶紫水晶招财手串手链", 35.0, 180, "天然紫水晶 弹力绳", "紫色", "招桃花 原创设计"),
        ("法式复古山茶花胸针女别针", 15.9, 300, "锌合金 烤漆", "金色 杏色", "山茶花造型 精致典雅"),
    ],
    "母婴玩具": [
        ("大颗粒积木早教益智玩具100粒", 79.0, 300, "ABS环保 BPA-free", "多彩", "大颗粒防吞咽 CE认证"),
        ("婴儿安抚巾毛绒玩偶可啃咬", 35.0, 200, "短毛绒 食品级硅胶", "粉色 蓝色", "可啃咬 安抚哄睡 不掉毛"),
        ("儿童保温杯316不锈钢吸管杯", 69.0, 400, "316不锈钢 PP", "粉色 蓝色", "316内胆 防摔 吸管两用"),
        ("婴儿爬行垫XPE加厚2cm双面折叠", 129.0, 100, "XPE PE覆膜", "双面图案", "防水防滑 缓冲减震 易收纳"),
        ("儿童平衡车无脚踏滑步车2-5岁", 199.0, 80, "碳钢 发泡轮", "红色 蓝色", "锻炼平衡 轻便防滑"),
    ],
    "宠物用品": [
        ("猫抓板瓦楞纸大号猫窝耐磨", 29.9, 350, "瓦楞纸", "牛皮纸色", "不掉屑 耐抓 可当猫窝"),
        ("狗狗磨牙耐咬发声玩具球互动", 19.9, 500, "TPR橡胶", "蓝色 橙色", "耐咬 发声 洁齿 互动"),
        ("宠物外出背包猫狗太空舱便携", 89.0, 120, "牛津布 PC透明窗", "黑色 粉色", "透气网窗 承重20斤"),
        ("狗窝秋冬保暖可拆洗猫狗通用", 49.0, 250, "水晶绒 PP棉", "灰色 棕色", "可拆洗 防滑底 保暖"),
        ("猫砂盆全封闭防外溅顶入式", 79.0, 180, "PP塑料", "白色 灰色", "防外溅 除臭 易清洗"),
        ("宠物自动饮水机循环过滤猫狗", 99.0, 150, "ABS食品级 PP", "白色", "静音水泵 大容量2.5L"),
    ],
    "厨房餐饮": [
        ("麦饭石不粘锅炒锅电磁炉通用32cm", 109.0, 200, "铝合金锅体 麦饭石涂层", "黑色 灰色", "不粘 轻油烟 加深锅型"),
        ("不锈钢可伸缩蒸笼蒸鱼蒸包子", 39.0, 300, "304不锈钢", "银色", "可调尺寸 多层蒸煮"),
        ("高硼硅玻璃饭盒分格便当微波炉", 49.0, 350, "高硼硅玻璃 PP盖", "透明", "分格不漏味 可微波 640ml*3"),
        ("不锈钢保温杯大容量1000ml户外", 59.0, 400, "316不锈钢内胆", "黑色 白色 绿色", "24小时保温 车载可用"),
        ("硅胶冰格模具带盖制冰盒64格", 19.9, 600, "食品级硅胶", "多色", "轻松取冰 带盖防串味"),
    ],
    "办公文具": [
        ("A5活页笔记本简约商务手账日记", 25.0, 500, "PU封面 道林纸", "棕色 黑色", "可替换内芯 180度平摊"),
        ("大容量笔袋ins风多层文具盒", 19.9, 400, "帆布", "米色 灰色 粉色", "大容量 多层分区 可洗"),
        ("手账贴纸和纸胶带套装200枚", 9.9, 800, "和纸", "多种", "不残留胶 可撕可贴 复古"),
        ("学生桌面收纳盒ins风笔筒", 29.9, 300, "PP塑料", "白色", "多格分类 不占空间"),
    ],
}


def _bold_header(ws, row, count):
    for c in range(1, count + 1):
        ws.cell(row, c).font = Font(bold=True)


# ── 格式 1: 1688 官方后台导出 ──
def fmt1(wb, prods, out_dir):
    ws = wb.active; ws.title = "1688后台导出"
    h = ["商品ID", "商品标题", "SKU编码", "价格(元)", "库存", "类目", "描述", "主图链接"]
    ws.append(h); _bold_header(ws, 1, len(h))
    pid = 1000001
    for cat, items in prods.items():
        for t, p, s, m, c, d in items:
            ws.append([f"PID{pid}", t, f"SKU{pid}", p, s, cat, d, f"https://img.1688.com/{pid}.jpg"])
            pid += 1


# ── 格式 2: 店小秘 ERP 导出 ──
def fmt2(wb, prods, out_dir):
    ws = wb.active; ws.title = "店小秘导出"
    h = ["SKU", "中文标题", "英文标题", "售价CNY", "库存", "重量(g)", "尺寸", "颜色", "材质", "图片URL"]
    ws.append(h); _bold_header(ws, 1, len(h))
    for cat, items in prods.items():
        for t, p, s, m, c, d in items:
            sz = random.choice(["S,M,L", "均码", "30x20cm", "10片", "单支", ""])
            ws.append([f"ERP-{hash(t)%100000:05d}", t, "", p, s, random.choice([200,350,500]), sz, c, m, f"https://img.erp.com/{hash(t)%1000}.jpg"])


# ── 格式 3: 工厂报价表（合并单元格 + 品类名）──
def fmt3(wb, prods, out_dir):
    ws = wb.active; ws.title = "工厂报价表"
    r = 1
    for cat, items in prods.items():
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
        c0 = ws.cell(r, 1, cat)
        c0.font = Font(bold=True, size=14)
        c0.alignment = Alignment(horizontal="center")
        r += 1
        for ci, h in enumerate(["型号", "产品名称", "材质", "颜色", "批发价"], 1):
            ws.cell(r, ci, h).font = Font(bold=True)
        r += 1
        for t, p, s, m, c, d in items[:4]:
            ws.append([f"LV{100+r:03d}", t, m, c, p])
            r += 1
        r += 1


# ── 格式 4: Shopify CSV ──
def fmt4(wb, prods, out_dir):
    path = out_dir / "4_shopify_export.csv"
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["Handle","Title","Body (HTML)","Vendor","Type","Tags","Published","Option1 Name","Option1 Value","Variant SKU","Variant Price","Variant Inventory Qty","Image Src"])
        for cat, items in prods.items():
            for t, p, s, m, c, d in items:
                hdl = t[:40].replace(" ", "-").replace("/", "-").lower()
                w.writerow([hdl, t, f"<p>{d}</p>", "1688-Supplier", cat, f"{cat}, {m}", "TRUE", "Color", c, f"SKU-{hash(t)%99999:05d}", p, s, f"https://img.shopify.com/{hash(t)%1000}.jpg"])
    return path


# ── 格式 5: 1688 插件采集 ──
def fmt5(wb, prods, out_dir):
    ws = wb.active; ws.title = "插件采集"
    h = ["商品链接", "商品标题", "价格(元)", "月销量", "评价数", "掌柜昵称", "店铺名", "所在地", "图片链接", "SKU列表"]
    ws.append(h); _bold_header(ws, 1, len(h))
    stores = ["义乌某某电子商务", "广州潮流饰品", "杭州某某服装", "深圳电子科技"]
    for cat, items in prods.items():
        for t, p, s, m, c, d in items:
            ws.append([f"https://detail.1688.com/offer/{random.randint(7000000000,7999999999)}.html", t, p, random.randint(50,5000), random.randint(5,500), "掌柜"+str(random.randint(100,999)), random.choice(stores), random.choice(["浙江金华","广东广州","广东深圳","浙江义乌"]), f"https://img.jpg", f"颜色:{c} | 材质:{m}"])


# ── 格式 6: 微信手工表 ──
def fmt6(wb, prods, out_dir):
    ws = wb.active; ws.title = "微信手工表"
    h = ["商品名", "单价", "备注"]
    ws.append(h); _bold_header(ws, 1, len(h))
    for cat, items in prods.items():
        for t, p, s, m, c, d in items[:3]:
            ws.append([t, p, f"{m} {c}" if m or c else cat])


# ── 格式 7: 1688 分销代发后台 ──
def fmt7(wb, prods, out_dir):
    ws = wb.active; ws.title = "分销代发"
    h = ["产品编码", "商品名称", "类目", "批发价(元)", "建议零售价(元)", "库存", "主图", "规格属性", "运费模板"]
    ws.append(h); _bold_header(ws, 1, len(h))
    for cat, items in prods.items():
        for t, p, s, m, c, d in items:
            ws.append([f"FX-{hash(t)%99999:05d}", t, cat, p, round(p*2.5,1), s, f"https://img.fx.com/{hash(t)%1000}.jpg", f"颜色:{c} | 材质:{m}", "全国包邮"])


# ── 格式 8: 多 Sheet 杂乱格式 ──
def fmt8(wb, prods, out_dir):
    sheets = [
        ("美容", ["产品", "规格", "成分", "价格", "颜色"]),
        ("饰品", ["名称", "材质（主体）", "镀层", "重量g", "批发价", "颜色分类"]),
        ("母婴", ["商品名", "规格尺寸", "材质", "价格"]),
        ("宠物", ["商品", "材质", "颜色", "订货价", "库存量"]),
    ]
    first = True
    for sname, headers in sheets:
        ws = wb.active if first else wb.create_sheet(title=sname)
        first = False; ws.title = sname
        ws.append(headers); _bold_header(ws, 1, len(headers))
        matched = None
        for cat in prods:
            if sname[:2] in cat: matched = cat; break
        if not matched: continue
        for t, p, s, m, c, d in prods[matched][:5]:
            row = []
            for h in headers:
                if "产" in h or "商" in h or "名" in h: row.append(t)
                elif "材" in h or "成" in h or "镀" in h: row.append(m)
                elif "颜" in h: row.append(c)
                elif "价" in h: row.append(p)
                elif "规" in h or "尺" in h: row.append(d[:30])
                elif "重" in h: row.append(random.choice([50,80,120,200]))
                elif "库" in h: row.append(s)
                else: row.append("")
            ws.append(row)


def generate_all(out_dir: Path):
    out_dir.mkdir(parents=True, exist_ok=True)
    prods = _PRODUCTS
    random.seed(42)

    generators = [
        ("1_1688官方后台导出.xlsx", fmt1),
        ("2_店小秘ERP导出.xlsx", fmt2),
        ("3_工厂报价表.xlsx", fmt3),
        ("5_1688插件采集.xlsx", fmt5),
        ("6_微信手工表.xlsx", fmt6),
        ("7_1688分销代发后台.xlsx", fmt7),
        ("8_多Sheet杂乱格式.xlsx", fmt8),
    ]

    for fname, gen in generators:
        wb = openpyxl.Workbook()
        gen(wb, prods, out_dir)
        wb.save(out_dir / fname)
        print(f"  {fname}")

    csv_path = fmt4(None, prods, out_dir)
    print(f"  {csv_path.name}")
    print(f"\n共 {len(generators)+1} 个测试文件 → {out_dir}/")


if __name__ == "__main__":
    import argparse
    p = argparse.ArgumentParser()
    p.add_argument("--dir", default=str(OUT_DIR))
    args = p.parse_args()
    generate_all(Path(args.dir))
