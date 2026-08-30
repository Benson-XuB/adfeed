"""FastAPI 路由 — 上传、处理、导出、账户"""
import os, json, uuid, tempfile, asyncio, time, logging
from pathlib import Path
from datetime import datetime, timezone
from typing import Optional

from fastapi import FastAPI, UploadFile, File, Form, HTTPException, Depends, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse, FileResponse, RedirectResponse, HTMLResponse, Response
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from pydantic import BaseModel, EmailStr

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
logger = logging.getLogger("adfeed-api")

from .db import (
    User, Job, get_user, list_jobs, get_job, create_job, update_job,
    get_user_by_email, increment_quota,
    get_shopify_connection, delete_shopify_connection,
)
from .auth import (
    create_jwt, decode_jwt, google_login, send_magic_link_email,
    verify_magic_link_and_login, get_google_auth_url,
)
from . import config
from .config import DATA_DIR, OUTPUT_DIR
from .shopify_auth import require_store
from .store_db import Store as StoreModel


# ── App ──

app = FastAPI(title="AdFeed AI", version="0.3.0", request_max_size=200 * 1024 * 1024)

from .platforms.google.router import router as google_platform_router
from .platforms.meta.router import router as meta_platform_router
from .platforms.tiktok.router import router as tiktok_platform_router

app.include_router(google_platform_router)
app.include_router(meta_platform_router)
app.include_router(tiktok_platform_router)


# Embedded admin + CLI tunnels call the API from HTTPS origins (not localhost UI).
# Admin UI extensions run on extensions.shopifycdn.com (CORS preflight Origin).
# Admin UI extensions: extensions.shopifycdn.com
# iframe React Router App Home: Shopify CLI tunnel + local Vite ports
_cors_origins = [
    "http://localhost:3000",
    "http://localhost:3001",
    "http://localhost:5173",
    "http://localhost:3458",
    "https://deltfu.com",
    "https://admin.shopify.com",
    "https://extensions.shopifycdn.com",
    "https://cdn.shopify.com",
    os.getenv("FRONTEND_URL", "http://localhost:3000"),
]
app.add_middleware(
    CORSMiddleware,
    allow_origins=[o for o in _cors_origins if o],
    allow_origin_regex=r"https://(.*\.)?(myshopify\.com|trycloudflare\.com|ngrok-free\.app|ngrok\.io|shopifypreview\.com|shopifycdn\.com|shopify\.com)",
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

security = HTTPBearer(auto_error=False)


@app.middleware("http")
async def log_requests(request: Request, call_next):
    start = time.time()
    response = await call_next(request)
    elapsed_ms = int((time.time() - start) * 1000)
    logger.info(f"{request.method} {request.url.path} → {response.status_code} ({elapsed_ms}ms)")
    return response


@app.get("/api/health")
async def health():
    return {"status": "ok", "ts": datetime.now(timezone.utc).isoformat()}


@app.get("/")
async def app_root():
    """Shopify embeds application_url. Merchant UI is React Router iframe App Home.
    Never return JSON 404 here — Admin shows it as a broken app page.
    """
    return HTMLResponse(
        """<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="utf-8"/>
  <meta name="viewport" content="width=device-width, initial-scale=1"/>
  <title>AdFeed AI</title>
  <style>
    body{font-family:-apple-system,BlinkMacSystemFont,Segoe UI,Helvetica,Arial,sans-serif;
      margin:0;padding:48px 24px;background:#f6f6f7;color:#202223;text-align:center}
    .card{max-width:520px;margin:0 auto;background:#fff;border:1px solid #e1e3e5;
      border-radius:12px;padding:28px 24px}
    h1{font-size:22px;margin:0 0 8px}
    p{color:#6d7175;line-height:1.5;margin:8px 0}
    code{background:#f1f2f3;padding:2px 6px;border-radius:4px;font-size:13px}
  </style>
</head>
<body>
  <div class="card">
    <h1>AdFeed AI</h1>
    <p>API is running. The merchant UI is the <strong>embedded iframe App Home</strong>
       (React Router) inside Shopify Admin — not this page.</p>
    <p>If you see this instead of Home/Plans, <code>application_url</code> still points at
       the API host. Point it at the web app (see <code>shopify.app.toml.prod-backup</code>)
       and reopen <strong>Apps → AdFeed AI</strong>.</p>
    <p><code>/api/health</code> OK</p>
  </div>
</body>
</html>""",
        status_code=200,
    )


@app.get("/privacy")
@app.get("/api/privacy")
async def privacy_policy():
    from .legal_pages import privacy_html
    return HTMLResponse(privacy_html())


@app.get("/support")
@app.get("/api/support")
async def support_page():
    from .legal_pages import support_html
    return HTMLResponse(support_html())


# ═══════════════ Shopify App APIs (session required) ═══════════════

@app.get("/api/app/billing/status")
async def app_billing_status(store: StoreModel = Depends(require_store)):
    """Return plan + quota for the authenticated shop."""
    return {
        "store_id": store.id,
        "shop_domain": store.shopify_domain,
        "shop_name": store.shop_name,
        "plan": store.plan,
        "billing_status": store.billing_status,
        "quota_total": store.quota_total,
        "quota_used": store.quota_used,
        "quota_remaining": store.quota_remaining,
        "subscription_id": store.subscription_id,
    }


class BillingSubscribeBody(BaseModel):
    plan: str = "starter"
    return_url: Optional[str] = None


@app.post("/api/app/billing/subscribe")
async def app_billing_subscribe(
    body: BillingSubscribeBody,
    store: StoreModel = Depends(require_store),
):
    """Create Shopify recurring subscription; return confirmation URL."""
    from .shopify_billing import create_app_subscription

    return_url = body.return_url or os.getenv(
        "ADFEED_BILLING_RETURN_URL",
        "https://deltfu.com/api/app/billing/return",
    )
    try:
        result = await create_app_subscription(
            store=store,
            plan=body.plan,
            return_url=return_url,
        )
    except ValueError as e:
        raise HTTPException(400, str(e)) from e
    except Exception as e:
        logger.error(f"Billing subscribe failed: {e}")
        msg = str(e)
        # Surface Shopify Partner/distribution misconfig clearly (UI often shows "Failed to fetch" on 502).
        if "owned by a Shop" in msg or "migrated to the Shopify partners" in msg.lower():
            raise HTTPException(
                400,
                "Shopify refused to create a subscription. In Partner Dashboard → this app → "
                "Distribution, choose Custom or Public first. Without Distribution, Billing API "
                "returns “owned by a Shop”. Then reinstall the app on this store and try again.",
            ) from e
        raise HTTPException(400, f"Shopify billing error: {e}") from e
    return result


@app.get("/api/app/billing/return")
async def app_billing_return(shop: Optional[str] = None, charge_id: Optional[str] = None):
    """After Shopify charge approval, send merchant back into Admin."""
    raw = (shop or "").strip().replace("https://", "").replace("http://", "")
    host = raw.replace(".myshopify.com", "").split("/")[0]
    app_key = config.SHOPIFY_CLIENT_ID or os.getenv("SHOPIFY_CLIENT_ID", "")
    if host and app_key:
        return RedirectResponse(f"https://admin.shopify.com/store/{host}/apps/{app_key}")
    if host:
        return RedirectResponse(f"https://admin.shopify.com/store/{host}")
    return RedirectResponse("https://admin.shopify.com")


@app.post("/api/webhooks/shopify/app_subscriptions_update")
async def webhook_app_subscriptions_update(request: Request):
    """APP_SUBSCRIPTIONS_UPDATE — sync plan + quota_total."""
    from .shopify_webhooks import verify_shopify_hmac
    from .shopify_billing import apply_subscription_webhook

    raw = await request.body()
    if not verify_shopify_hmac(raw, request.headers.get("X-Shopify-Hmac-Sha256", "")):
        # Allow unsigned in tests when secret empty / ADFEED_WEBHOOK_SKIP_HMAC
        if os.getenv("ADFEED_WEBHOOK_SKIP_HMAC", "").lower() not in ("1", "true", "yes"):
            if config.SHOPIFY_CLIENT_SECRET:
                raise HTTPException(401, "Invalid webhook HMAC")

    try:
        payload = json.loads(raw.decode("utf-8") or "{}")
    except json.JSONDecodeError as e:
        raise HTTPException(400, "Invalid JSON") from e

    # Prefer shop domain from header when present
    shop_header = request.headers.get("X-Shopify-Shop-Domain", "")
    if shop_header and "shop_domain" not in payload:
        payload["shop_domain"] = shop_header

    store = apply_subscription_webhook(payload)
    return {"ok": True, "store_id": store.id if store else None}


@app.get("/api/app/products")
async def app_list_products(store: StoreModel = Depends(require_store)):
    """List products for the shop (live Shopify pull when token present)."""
    from . import store_db
    from .shopify_client import fetch_shopify_products

    products = []
    source = "cache"

    if store.access_token:
        try:
            data = await fetch_shopify_products(
                shop_domain=store.shopify_domain,
                access_token=store.access_token,
                limit=250,
                lite=True,
            )
            products = [
                {
                    "id": str(p.get("shopify_id") or p.get("SKU") or ""),
                    "title": p.get("标题") or p.get("title") or "",
                    "image_url": p.get("图片链接") or p.get("image_url") or "",
                    "price": p.get("价格") or p.get("price") or 0,
                    "status": "active",
                    "need_color": bool(p.get("need_color")),
                    "need_size": bool(p.get("need_size")),
                    "product_type": p.get("product_type") or p.get("分类") or "",
                    "variant_count": int(p.get("variant_count") or 0),
                }
                for p in data.get("products", [])
            ]
            source = "shopify"
        except Exception as e:
            logger.warning(f"Shopify product pull failed: {e}")

    if not products:
        from .product_attr_check import check_store_product_attrs

        cached = store_db.get_store_products(store.id)
        products = []
        for p in cached:
            gaps = check_store_product_attrs(p, store_db.get_product_variants(p.id))
            products.append({
                "id": p.shopify_product_id or p.id,
                "title": p.title,
                "image_url": p.image_url or "",
                "price": 0,
                "status": p.status,
                "need_color": gaps["need_color"],
                "need_size": gaps["need_size"],
                "product_type": p.product_type or "",
                "variant_count": 0,
            })
        source = "cache"

    return {
        "store_id": store.id,
        "shop_domain": store.shopify_domain,
        "source": source,
        "products": products,
        "count": len(products),
    }


class AppGenerateBody(BaseModel):
    product_ids: list[str]
    platforms: list[str] = ["google"]
    languages: list[str] = ["US"]
    # True: optimize only product_ids, then write XML = (ids already in durable
    # feeds) ∪ product_ids. False (default): XML contains only product_ids.
    merge: bool = False


@app.post("/api/app/bootstrap")
async def app_bootstrap(
    request: Request,
    store: StoreModel = Depends(require_store),
):
    """Exchange session JWT → offline Admin token and persist on stores."""
    from . import store_db
    from .store_sync import ensure_store_access_token

    auth = request.headers.get("Authorization", "")
    session_token = auth[7:].strip() if auth.lower().startswith("bearer ") else ""
    if not session_token:
        raise HTTPException(401, "Missing session token")

    # Re-exchange when token is missing/stale/non-expiring (Admin API 403).
    updated = await ensure_store_access_token(store, session_token, force=False)
    # Refresh shop currency so App can guide market selection
    if updated.access_token:
        from .store_sync import refresh_store_currency_from_shopify
        updated = refresh_store_currency_from_shopify(updated)
    has_token = bool(updated.access_token)
    return {
        "ok": has_token,
        "store_id": updated.id,
        "shop_domain": updated.shopify_domain,
        "shop_name": updated.shop_name,
        "has_access_token": has_token,
        "default_currency": (updated.default_currency or "USD").upper(),
        "plan": updated.plan,
        "quota_remaining": updated.quota_remaining,
        "message": (
            "Offline token ready"
            if has_token
            else "Token exchange failed — check SHOPIFY_CLIENT_ID/SECRET and app install"
        ),
    }


@app.get("/api/app/connection")
async def app_connection(store: StoreModel = Depends(require_store)):
    return {
        "store_id": store.id,
        "shop_domain": store.shopify_domain,
        "shop_name": store.shop_name,
        "has_access_token": bool(store.access_token),
        "site_url": store.site_url,
        "default_brand": store.default_brand or "",
        "default_currency": (store.default_currency or "USD").upper(),
        "status": store.status,
        "plan": store.plan,
        "quota_remaining": store.quota_remaining,
    }


@app.get("/api/app/compatible-markets")
async def app_compatible_markets(store: StoreModel = Depends(require_store)):
    """Batch GREEN countries for market dropdown (shop currency + Markets + preflight)."""
    from .compatible_markets import list_compatible_markets

    return list_compatible_markets(
        store_id=store.id,
        shop_domain=store.shopify_domain or "",
        access_token=store.access_token,
        shop_currency=store.default_currency or "USD",
    )


@app.get("/api/app/market-ready")
async def app_market_ready(
    country: str = "US",
    store: StoreModel = Depends(require_store),
):
    """Can this shop emit a feed for ``country`` without inventing FX?"""
    from . import store_db
    from .market_pricing import (
        PreflightStatus,
        expected_currency_for_country,
        preflight_country,
    )
    from .shopify_markets import fetch_contextual_pricing

    cu = (country or "US").strip().upper()
    shop_ccy = (store.default_currency or "USD").strip().upper()
    vids: list[str] = []
    for p in store_db.get_store_products(store.id)[:12]:
        for v in store_db.get_product_variants(p.id):
            vid = (v.shopify_variant_id or "").strip()
            if vid.isdigit():
                vids.append(vid)
            if len(vids) >= 5:
                break
        if len(vids) >= 5:
            break

    sample = None
    if store.access_token and store.shopify_domain and vids:
        fetched = fetch_contextual_pricing(
            store.shopify_domain, store.access_token, vids, cu,
        )
        if fetched:
            first = next(iter(fetched.values()))
            sample = {cu: first}

    pf = preflight_country(shop_currency=shop_ccy, country=cu, sample_presentment=sample)
    return {
        "country": cu,
        "ready": pf.status == PreflightStatus.GREEN,
        "shop_currency": shop_ccy,
        "expected_currency": expected_currency_for_country(cu),
        "message": pf.message,
    }


@app.patch("/api/app/store/brand")
async def app_update_store_brand(
    request: Request,
    store: StoreModel = Depends(require_store),
):
    """Confirm advertising brand written to g:brand (field contract owner: store)."""
    from . import store_db

    body = await request.json()
    brand = (body.get("default_brand") or "").strip()
    if len(brand) > 70:
        raise HTTPException(400, "Brand name too long (max 70 characters)")
    store_db.update_store(store.id, default_brand=brand or None)
    updated = store_db.get_store(store.id)
    return {
        "store_id": updated.id,
        "default_brand": updated.default_brand or "",
        "shop_name": updated.shop_name,
    }


@app.get("/api/app/store/compliance")
async def app_store_compliance(
    request: Request,
    store: StoreModel = Depends(require_store),
    countries: str = "US",
):
    """Lite storefront compliance scan (policies + HTTPS + contact + currency hint)."""
    from .store_sync import ensure_store_access_token
    from .store_compliance import diagnose_store_compliance

    auth = request.headers.get("Authorization", "")
    session_token = auth[7:].strip() if auth.lower().startswith("bearer ") else ""
    if session_token and not store.access_token:
        store = await ensure_store_access_token(store, session_token)

    if not store.access_token:
        raise HTTPException(
            409,
            "Store is not connected to the Admin API yet. Open the app to finish bootstrap.",
        )

    cu_list = [c.strip().upper() for c in (countries or "US").split(",") if c.strip()]
    if not cu_list:
        cu_list = ["US"]

    report = diagnose_store_compliance(
        shop_domain=store.shopify_domain,
        access_token=store.access_token,
        site_url=store.site_url or "",
        shop_currency=store.default_currency or "USD",
        countries=cu_list,
    )
    return report.to_dict()


@app.post("/api/app/generate")
async def app_generate(
    request: Request,
    body: AppGenerateBody,
    store: StoreModel = Depends(require_store),
):
    """Start layered optimize + feed generate job (SKU×platform×language quota)."""
    from . import store_db
    from .quota import estimate_cost, assert_quota_available
    from .store_sync import ensure_store_access_token, sync_products_for_generate, normalize_shopify_product_id

    platforms = [p.lower() for p in (body.platforms or ["google"])] or ["google"]
    languages = [l.upper() for l in (body.languages or ["US"])] or ["US"]
    product_ids = [normalize_shopify_product_id(x) for x in (body.product_ids or [])]
    product_ids = [p for p in product_ids if p]
    if not product_ids:
        raise HTTPException(400, "product_ids required")

    # Ensure offline token before sync
    auth = request.headers.get("Authorization", "")
    session_token = auth[7:].strip() if auth.lower().startswith("bearer ") else ""
    if session_token and not store.access_token:
        store = await ensure_store_access_token(store, session_token)

    if not store.access_token:
        raise HTTPException(
            409,
            "No Admin API token yet. Open the app (bootstrap runs automatically on load).",
        )

    # Apparel feeds need g:brand — block until merchant confirms advertising brand
    store = store_db.get_store(store.id) or store
    if not (store.default_brand or "").strip():
        raise HTTPException(
            409,
            "Confirm an ad brand in the app before generating a feed. Apparel on Google needs "
            "brand; an empty brand often fails as Missing brand.",
        )

    # Sync selected Shopify products (+ variants) into store_db
    try:
        internal_ids = await sync_products_for_generate(store, product_ids)
    except Exception as e:
        logger.exception("product sync failed")
        raise HTTPException(502, f"Product sync failed: {e}") from e

    if not internal_ids:
        raise HTTPException(
            404,
            "Could not sync any products locally. Check product IDs and read_products scope.",
        )

    cost = estimate_cost(len(internal_ids), platforms, languages)
    store = store_db.get_store(store.id) or store
    assert_quota_available(store, cost)

    job = store_db.create_store_job(
        store_id=store.id,
        platforms=platforms,
        languages=languages,
        product_ids=internal_ids,
        total_units=cost,
    )
    store_db.update_store_job(job.id, status="processing")

    import threading
    store_id = store.id

    merge = bool(getattr(body, "merge", False))

    def _run():
        from .pipeline import optimize_layered, generate_feed_for_store
        from .feed_preview import internal_ids_in_durable_feeds
        try:
            def progress(done, total):
                store_db.update_store_job(job.id, done_units=done, total_units=total)

            opt = optimize_layered(
                store_id=store_id,
                product_ids=internal_ids,
                platforms=platforms,
                languages=languages,
                job_id=job.id,
                progress_callback=progress,
            )
            write_ids = list(internal_ids)
            if merge:
                existing = internal_ids_in_durable_feeds(
                    store_id,
                    platforms=platforms,
                    countries=languages,
                )
                write_ids = list(existing | set(internal_ids))
            feeds = generate_feed_for_store(
                store_id=store_id,
                countries=languages,
                platforms=platforms,
                product_ids=write_ids,
            )
            result = {
                "optimize": {
                    "ok_units": opt.get("ok_units", 0),
                    "fail_units": opt.get("fail_units", 0),
                    "assets_written": opt.get("assets_written", 0),
                },
                "feeds": feeds.get("feed_urls", []),
                "blocked_countries": feeds.get("blocked_countries", []),
                "quality_report": feeds.get("quality_report"),
                "synced_products": len(internal_ids),
                "write_products": len(write_ids),
                "merge": merge,
                "message": feeds.get("message"),
            }
            store_db.update_store_job(
                job.id,
                status="completed",
                ok_units=opt.get("ok_units", 0),
                fail_units=opt.get("fail_units", 0),
                done_units=opt.get("ok_units", 0) + opt.get("fail_units", 0),
                result_json=json.dumps(result, ensure_ascii=False),
            )
        except Exception as e:
            logger.exception("app generate failed")
            store_db.update_store_job(job.id, status="failed", error_msg=str(e))

    threading.Thread(target=_run, daemon=True).start()
    return {
        "job_id": job.id,
        "status": "processing",
        "estimate": cost,
        "platforms": platforms,
        "languages": languages,
        "product_count": len(internal_ids),
        "synced_products": len(internal_ids),
    }


@app.get("/api/app/jobs/{job_id}")
async def app_job_status(job_id: str, store: StoreModel = Depends(require_store)):
    from . import store_db
    job = store_db.get_store_job(job_id)
    if not job or job.store_id != store.id:
        raise HTTPException(404, "Job not found")
    result = None
    if job.result_json:
        try:
            result = json.loads(job.result_json)
        except json.JSONDecodeError:
            result = None
    return {
        "job_id": job.id,
        "status": job.status,
        "total_units": job.total_units,
        "done_units": job.done_units,
        "ok_units": job.ok_units,
        "fail_units": job.fail_units,
        "error_msg": job.error_msg,
        "result": result,
        "platforms": json.loads(job.platforms) if job.platforms else [],
        "languages": json.loads(job.languages) if job.languages else [],
    }


class BulkPatchItem(BaseModel):
    sku: str
    color: Optional[str] = None
    size: Optional[str] = None


class BulkPatchBody(BaseModel):
    patches: list[BulkPatchItem]
    platforms: list[str] = ["google"]
    languages: list[str] = ["US"]
    regenerate: bool = True
    shopify_product_id: Optional[str] = None


@app.post("/api/app/quality/bulk_patch")
async def app_quality_bulk_patch(
    body: BulkPatchBody,
    store: StoreModel = Depends(require_store),
):
    """Batch-confirm Multicolor / One Size (etc.) → DB + optional feed regenerate."""
    from .quality_bulk import bulk_patch_and_regen

    if not body.patches:
        raise HTTPException(400, "patches required")

    try:
        result = bulk_patch_and_regen(
            store.id,
            [p.model_dump() for p in body.patches],
            platforms=body.platforms,
            languages=body.languages,
            regenerate=body.regenerate,
            shopify_product_id=body.shopify_product_id,
        )
    except Exception as e:
        logger.exception("bulk_patch failed")
        raise HTTPException(500, f"Bulk patch failed: {e}") from e

    return result


class ShopifyVariantPatchBody(BaseModel):
    shopify_product_id: str
    patches: list[BulkPatchItem]


@app.post("/api/app/products/shopify_variant_patch")
async def app_shopify_variant_patch(
    body: ShopifyVariantPatchBody,
    store: StoreModel = Depends(require_store),
):
    """Write color/size to Shopify variant options, then sync store_db."""
    import asyncio

    from .product_attr_check import check_shopify_product_attrs
    from .shopify_variant_patch import patch_shopify_variant_attrs
    from .store_sync import fetch_raw_product, upsert_raw_shopify_product

    if not body.patches:
        raise HTTPException(400, "patches required")
    if not store.access_token:
        raise HTTPException(401, "Store not connected — reopen the App to authorize")

    pid = str(body.shopify_product_id or "").strip()
    if not pid:
        raise HTTPException(400, "shopify_product_id required")

    patch_dump = [p.model_dump() for p in body.patches]
    logger.info(
        "shopify_variant_patch request store=%s product=%s patches=%s",
        store.id,
        pid,
        patch_dump,
    )

    try:
        result = await asyncio.to_thread(
            patch_shopify_variant_attrs,
            store.shopify_domain,
            store.access_token,
            pid,
            patch_dump,
        )
    except Exception as e:
        logger.exception("shopify_variant_patch failed")
        raise HTTPException(500, f"Shopify write failed: {e}") from e

    logger.info(
        "shopify_variant_patch result store=%s product=%s updated=%s message=%s debug=%s",
        store.id,
        pid,
        result.get("updated"),
        result.get("message"),
        result.get("debug"),
    )

    if not result.get("updated"):
        msg = result.get("message") or "No variants were updated"
        code = 403 if result.get("need_reauth") else 400
        raise HTTPException(code, msg)

    raw = await fetch_raw_product(store.shopify_domain, store.access_token, pid)
    gaps: dict = {}
    if raw:
        upsert_raw_shopify_product(store.id, raw)
        gaps = check_shopify_product_attrs(raw)

    return {
        **result,
        "need_color": bool(gaps.get("need_color")),
        "need_size": bool(gaps.get("need_size")),
        "gaps": gaps,
    }


class ImagePatchItem(BaseModel):
    sku: str
    image_url: str


class ImagePatchBody(BaseModel):
    patches: list[ImagePatchItem]
    platforms: list[str] = ["google"]
    languages: list[str] = ["US"]
    regenerate: bool = True


@app.get("/api/app/feed-images")
async def app_feed_images(
    sku: str,
    store: StoreModel = Depends(require_store),
):
    """Feed main-image candidates + recommendation for one SKU."""
    from .feed_image import get_feed_image_context

    sku = (sku or "").strip()
    if not sku:
        raise HTTPException(400, "sku required")
    ctx = get_feed_image_context(store.id, sku)
    if not ctx:
        raise HTTPException(404, f"SKU not found: {sku}")
    return ctx


@app.post("/api/app/quality/image_patch")
async def app_quality_image_patch(
    body: ImagePatchBody,
    store: StoreModel = Depends(require_store),
):
    """Set per-SKU feed_image_url override → optional feed regenerate."""
    from .feed_image import image_patch_and_regen

    if not body.patches:
        raise HTTPException(400, "patches required")
    try:
        result = image_patch_and_regen(
            store.id,
            [p.model_dump() for p in body.patches],
            platforms=body.platforms,
            languages=body.languages,
            regenerate=body.regenerate,
        )
    except Exception as e:
        logger.exception("image_patch failed")
        raise HTTPException(500, f"Image update failed: {e}") from e
    return result


@app.get("/api/app/feeds")
async def app_list_feeds(store: StoreModel = Depends(require_store)):
    """List feeds + last completed job quality_report (for honest KPI after refresh)."""
    import json as _json
    from . import store_db

    feeds = store_db.list_store_feeds(store.id)
    quality_report = None
    last_job = None
    job = store_db.get_latest_completed_store_job(store.id)
    if job:
        try:
            langs = _json.loads(job.languages) if job.languages else []
        except Exception:
            langs = []
        try:
            plats = _json.loads(job.platforms) if job.platforms else []
        except Exception:
            plats = []
        last_job = {
            "id": job.id,
            "languages": langs if isinstance(langs, list) else [],
            "platforms": plats if isinstance(plats, list) else [],
            "updated_at": job.updated_at,
        }
        if job.result_json:
            try:
                result = _json.loads(job.result_json)
                if isinstance(result, dict):
                    quality_report = result.get("quality_report")
            except Exception:
                quality_report = None

    from .config import PUBLIC_BASE_URL
    from .multi_platform_feeds import durable_feed_url

    return {
        "feeds": [
            {
                "platform": f.platform,
                "country": f.country,
                "language": f.country,
                "url": durable_feed_url(
                    PUBLIC_BASE_URL, store.id, f.platform, f.country,
                ),
                "csv_url": durable_feed_url(
                    PUBLIC_BASE_URL, store.id, f.platform, f.country,
                ).replace(".xml", ".csv"),
                "item_count": f.item_count,
                "updated_at": f.generated_at,
            }
            for f in feeds
        ],
        "quality_report": quality_report,
        "last_job": last_job,
        "store_id": store.id,
    }


class FeedRowPatchItem(BaseModel):
    sku: str
    title: Optional[str] = None
    color: Optional[str] = None
    size: Optional[str] = None
    image_url: Optional[str] = None


class FeedRowPatchBody(BaseModel):
    patches: list[FeedRowPatchItem]
    platforms: list[str] = ["google"]
    languages: list[str] = ["US"]
    regenerate: bool = True


class FeedRowDeleteBody(BaseModel):
    skus: list[str]
    platforms: list[str] = ["google"]
    languages: list[str] = ["US"]


def _latest_quality_for_store(store_id: str):
    import json as _json
    from . import store_db

    job = store_db.get_latest_completed_store_job(store_id)
    if not job or not job.result_json:
        return None
    try:
        result = _json.loads(job.result_json)
        if isinstance(result, dict):
            return result.get("quality_report")
    except Exception:
        return None
    return None


@app.get("/api/app/feeds/{platform}/{country}/preview")
async def app_feed_preview(
    platform: str,
    country: str,
    store: StoreModel = Depends(require_store),
    limit: int = 20,
    offset: int = 0,
    q: str = "",
    product_id: str = "",
):
    """Paginated feed item preview from the durable current file."""
    from . import store_db
    from .feed_preview import preview_feed_items
    from .config import PUBLIC_BASE_URL
    from .multi_platform_feeds import durable_feed_url

    plat = (platform or "google").lower()
    cu = (country or "US").upper()
    feed = store_db.get_store_feed(store.id, cu, plat)
    if not feed or not feed.file_path:
        raise HTTPException(404, "Feed not found — generate first")

    data = preview_feed_items(
        file_path=feed.file_path,
        platform=plat,
        limit=limit,
        offset=offset,
        q=q,
        quality_report=_latest_quality_for_store(store.id),
        product_id=product_id,
        store_id=store.id,
    )
    public_url = durable_feed_url(PUBLIC_BASE_URL, store.id, plat, cu)
    return {
        **data,
        "platform": plat,
        "country": cu,
        "url": public_url,
        "csv_url": public_url.replace(".xml", ".csv"),
        "item_count": feed.item_count,
        "updated_at": feed.generated_at,
        "store_id": store.id,
    }


@app.get("/api/app/feeds/{platform}/{country}/workbench")
async def app_feed_workbench(
    platform: str,
    country: str,
    store: StoreModel = Depends(require_store),
):
    """Product-row workbench: products + per-product feed status for current feed."""
    from . import store_db
    from .feed_preview import build_workbench_product_rows
    from .config import PUBLIC_BASE_URL
    from .multi_platform_feeds import durable_feed_url
    from .shopify_client import fetch_shopify_products

    plat = (platform or "google").lower()
    cu = (country or "US").upper()
    feed = store_db.get_store_feed(store.id, cu, plat)

    products: list[dict] = []
    if store.access_token:
        try:
            data = await fetch_shopify_products(
                shop_domain=store.shopify_domain,
                access_token=store.access_token,
                limit=250,
                lite=True,
            )
            products = [
                {
                    "id": str(p.get("shopify_id") or p.get("SKU") or ""),
                    "title": p.get("标题") or p.get("title") or "",
                    "image_url": p.get("图片链接") or p.get("image_url") or "",
                    "price": p.get("价格") or p.get("price") or 0,
                    "status": "active",
                    "need_color": bool(p.get("need_color")),
                    "need_size": bool(p.get("need_size")),
                    "product_type": p.get("product_type") or p.get("分类") or "",
                    "variant_count": int(p.get("variant_count") or 0),
                    "variant_skus": list(p.get("variant_skus") or []),
                }
                for p in data.get("products", [])
            ]
        except Exception as e:
            logger.warning("workbench shopify pull failed: %s", e)
    if not products:
        from .product_attr_check import check_store_product_attrs

        cached = store_db.get_store_products(store.id)
        products = []
        for p in cached:
            gaps = check_store_product_attrs(p, store_db.get_product_variants(p.id))
            variants = store_db.get_product_variants(p.id)
            products.append({
                "id": p.shopify_product_id or p.id,
                "title": p.title,
                "image_url": p.image_url or "",
                "price": 0,
                "status": p.status,
                "need_color": gaps["need_color"],
                "need_size": gaps["need_size"],
                "product_type": p.product_type or "",
                "variant_count": len(variants),
                "variant_skus": [v.sku for v in variants if v.sku],
            })

    qr = _latest_quality_for_store(store.id)
    from .product_attr_check import check_store_product_attrs

    # Prefer local DB SKUs + patched color/size when store variants exist.
    by_shopify: dict[str, list[str]] = {}
    store_products_by_shopify: dict[str, object] = {}
    for sp in store_db.get_store_products(store.id):
        key = str(sp.shopify_product_id or sp.id)
        by_shopify[key] = [
            v.sku for v in store_db.get_product_variants(sp.id) if v.sku
        ]
        store_products_by_shopify[key] = sp
    for row in products:
        pid = str(row.get("id") or "")
        if not row.get("variant_skus"):
            row["variant_skus"] = by_shopify.get(pid, [])
        sp = store_products_by_shopify.get(pid)
        if sp:
            variants = store_db.get_product_variants(sp.id)
            if variants:
                gaps = check_store_product_attrs(sp, variants)
                row["need_color"] = gaps["need_color"]
                row["need_size"] = gaps["need_size"]
                row["gaps_from_store_db"] = True

    rows = build_workbench_product_rows(
        store_id=store.id,
        file_path=feed.file_path if feed else "",
        platform=plat,
        products=products,
        quality_report=qr,
    )
    public_url = (
        durable_feed_url(PUBLIC_BASE_URL, store.id, plat, cu) if feed else ""
    )
    return {
        "platform": plat,
        "country": cu,
        "products": rows,
        "count": len(rows),
        "feed": {
            "url": public_url,
            "csv_url": public_url.replace(".xml", ".csv") if public_url else "",
            "item_count": feed.item_count if feed else 0,
            "updated_at": feed.generated_at if feed else None,
            "exists": bool(feed),
        },
        "quality_report": qr,
        "store_id": store.id,
    }


@app.get("/api/app/feeds/{platform}/{country}/download.csv")
async def app_feed_download_csv(
    platform: str,
    country: str,
    store: StoreModel = Depends(require_store),
):
    """Auth download of Google TSV derived from current XML."""
    from pathlib import Path
    from fastapi.responses import FileResponse
    from . import store_db
    from .feed_preview import write_google_tsv_from_xml

    plat = (platform or "google").lower()
    cu = (country or "US").upper()
    feed = store_db.get_store_feed(store.id, cu, plat)
    if not feed or not feed.file_path:
        raise HTTPException(404, "Feed not found")
    xml_path = Path(feed.file_path)
    if not xml_path.exists():
        raise HTTPException(404, "Feed file missing on disk")
    csv_path = xml_path.with_suffix(".csv")
    if plat == "google":
        write_google_tsv_from_xml(xml_path, csv_path)
    elif not csv_path.exists():
        raise HTTPException(404, "CSV not available for this platform")
    return FileResponse(
        csv_path,
        media_type="text/tab-separated-values",
        filename=f"{plat}_{cu.lower()}.csv",
    )


@app.get("/api/app/feeds/{platform}/{country}/snapshots")
async def app_feed_snapshots(
    platform: str,
    country: str,
    store: StoreModel = Depends(require_store),
):
    from .feed_snapshots import list_snapshots

    return {
        "snapshots": list_snapshots(store.id, platform, country),
        "platform": (platform or "google").lower(),
        "country": (country or "US").upper(),
    }


@app.post("/api/app/feeds/snapshots/{snapshot_id}/restore")
async def app_feed_snapshot_restore(
    snapshot_id: str,
    store: StoreModel = Depends(require_store),
):
    from .feed_snapshots import restore_snapshot

    try:
        return restore_snapshot(store.id, snapshot_id)
    except ValueError as e:
        raise HTTPException(404, str(e)) from e


@app.get("/api/app/feeds/snapshots/{snapshot_id}/download")
async def app_feed_snapshot_download(
    snapshot_id: str,
    store: StoreModel = Depends(require_store),
):
    from pathlib import Path
    from fastapi.responses import FileResponse
    from .feed_snapshots import get_snapshot

    snap = get_snapshot(store.id, snapshot_id)
    if not snap:
        raise HTTPException(404, "snapshot not found")
    path = Path(snap["file_path"])
    if not path.exists():
        raise HTTPException(404, "snapshot file missing")
    return FileResponse(path, filename=path.name)


@app.post("/api/app/feeds/row_patch")
async def app_feed_row_patch(
    body: FeedRowPatchBody,
    store: StoreModel = Depends(require_store),
):
    """Edit title/color/size/image on owner layers, optionally regenerate feed."""
    from .feed_row_edit import row_patch_and_regen

    if not body.patches:
        raise HTTPException(400, "patches required")
    try:
        return row_patch_and_regen(
            store.id,
            [p.model_dump() for p in body.patches],
            platforms=body.platforms,
            languages=body.languages,
            regenerate=body.regenerate,
        )
    except Exception as e:
        logger.exception("row_patch failed")
        raise HTTPException(500, f"Row edit failed: {e}") from e


@app.post("/api/app/feeds/row_delete")
async def app_feed_row_delete(
    body: FeedRowDeleteBody,
    store: StoreModel = Depends(require_store),
):
    """Remove SKU rows from durable feed (does not delete Shopify products)."""
    from .feed_row_edit import delete_feed_rows

    if not body.skus:
        raise HTTPException(400, "skus required")
    try:
        plat = (body.platforms or ["google"])[0].lower()
        cu = (body.languages or ["US"])[0].upper()
        return delete_feed_rows(
            store.id,
            body.skus,
            platform=plat,
            country=cu,
        )
    except Exception as e:
        logger.exception("row_delete failed")
        raise HTTPException(500, f"Row delete failed: {e}") from e


@app.post("/api/app/quota/estimate")
async def app_quota_estimate(body: AppGenerateBody, store: StoreModel = Depends(require_store)):
    from .quota import estimate_cost
    platforms = body.platforms or ["google"]
    languages = body.languages or ["US"]
    cost = estimate_cost(len(body.product_ids or []), platforms, languages)
    return {
        "estimate": cost,
        "quota_remaining": store.quota_remaining,
        "affordable": cost <= store.quota_remaining,
        "platforms": platforms,
        "languages": languages,
        "sku_count": len(body.product_ids or []),
    }


# ── Models ──

class MagicLinkRequest(BaseModel):
    email: EmailStr

class MagicLinkVerifyRequest(BaseModel):
    token: str

class GoogleCallbackRequest(BaseModel):
    code: str
    state: str = "login"


# ── Auth Dependency ──

async def current_user(
    request: Request,
    credentials: Optional[HTTPAuthorizationCredentials] = Depends(security),
) -> User:
    # 也支持 cookie
    token = None
    if credentials:
        token = credentials.credentials
    if not token:
        token = request.cookies.get("adfeed_token")
    if not token:
        raise HTTPException(status_code=401, detail="Not signed in")

    payload = decode_jwt(token)
    if not payload:
        raise HTTPException(status_code=401, detail="Token expired")

    user = get_user(payload["sub"])
    if not user:
        raise HTTPException(status_code=401, detail="User not found")
    return user


# ═══════════════ Auth Routes ═══════════════

@app.get("/api/auth/google/url")
async def google_auth_url():
    return {"url": get_google_auth_url()}


@app.post("/api/auth/google/callback")
async def google_callback(body: GoogleCallbackRequest):
    result = await google_login(body.code)
    if not result:
        raise HTTPException(400, "Google sign-in failed")

    user, token = result
    resp = JSONResponse({
        "token": token,
        "user": {
            "id": user.id, "email": user.email, "name": user.name,
            "avatar_url": user.avatar_url, "plan": user.plan,
            "quota_total": user.quota_total, "quota_used": user.quota_used,
        },
    })
    resp.set_cookie("adfeed_token", token, httponly=True, max_age=86400*7, samesite="lax")
    return resp


def _require_web_saas():
    """Web SaaS paths retired unless ADFEED_WEB_SAAS_ENABLED=true."""
    if not getattr(config, "WEB_SAAS_ENABLED", False):
        raise HTTPException(
            410,
            "Web SaaS auth/upload is retired. Install the Shopify App instead.",
        )


@app.post("/api/auth/magic-link")
async def request_magic_link(body: MagicLinkRequest):
    _require_web_saas()
    from .db import create_magic_link
    token = create_magic_link(body.email)
    await send_magic_link_email(body.email, token)
    return {"ok": True, "message": "Magic link sent to your email"}


@app.get("/api/auth/magic-link/verify")
async def verify_magic_link(token: str):
    _require_web_saas()
    result = verify_magic_link_and_login(token)
    if not result:
        raise HTTPException(400, "Link is invalid or expired")

    user, jwt_token = result
    resp = JSONResponse({
        "token": jwt_token,
        "user": {
            "id": user.id, "email": user.email, "name": user.name,
            "avatar_url": user.avatar_url, "plan": user.plan,
            "quota_total": user.quota_total, "quota_used": user.quota_used,
        },
    })
    resp.set_cookie("adfeed_token", jwt_token, httponly=True, max_age=86400*7, samesite="lax")
    return resp


@app.get("/api/auth/me")
async def me(user: User = Depends(current_user)):
    return {
        "id": user.id, "email": user.email, "name": user.name,
        "avatar_url": user.avatar_url, "plan": user.plan,
        "quota_total": user.quota_total, "quota_used": user.quota_used,
        "quota_remaining": user.quota_remaining,
    }


# ═══════════════ Upload ═══════════════

UPLOAD_DIR = DATA_DIR / "uploads"
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)

ALLOWED_EXTENSIONS = {".xlsx", ".xls", ".csv", ".txt"}


@app.post("/api/upload")
async def upload_file(
    file: UploadFile = File(...),
    countries: str = Form('["US"]'),
    user: User = Depends(current_user),
):
    _require_web_saas()
    # 校验扩展名
    ext = Path(file.filename).suffix.lower()
    if ext not in ALLOWED_EXTENSIONS:
        raise HTTPException(400, f"Unsupported file type: {ext}. Allowed: {', '.join(ALLOWED_EXTENSIONS)}")

    # 流式保存文件（不一次性加载 200MB 到内存）
    file_id = str(uuid.uuid4())
    safe_name = f"{file_id}{ext}"
    file_path = UPLOAD_DIR / safe_name

    import hashlib
    h = hashlib.md5()
    t1 = time.time()
    bytes_total = 0
    with open(file_path, "wb") as f:
        while chunk := await file.read(4 * 1024 * 1024):  # 4MB chunks
            f.write(chunk)
            h.update(chunk)
            bytes_total += len(chunk)
    t2 = time.time()
    file_hash = h.hexdigest()
    logger.info(f"upload_file write_disk: {bytes_total/1024/1024:.1f} MB in {(t2-t1)*1000:.0f}ms ({bytes_total/(t2-t1)/1024/1024:.1f} MB/s)")

    # 校验 country_mask 格式
    try:
        countries_parsed = json.loads(countries)
    except json.JSONDecodeError:
        raise HTTPException(400, 'Invalid countries format — use a JSON array, e.g. ["US","DE"]')

    # 创建任务，状态为 analyzing
    job = create_job(user.id, file.filename, json.dumps(countries_parsed), file_hash)

    # 后台线程分析文件（不阻塞上传响应）
    import threading
    threading.Thread(target=_analyze_upload, args=(job.id, str(file_path)), daemon=True).start()

    return {
        "job_id": job.id,
        "filename": file.filename,
        "countries": countries_parsed,
        "quota_remaining": user.quota_remaining,
        "quota_total": user.quota_total,
        "status": "analyzing",
    }


def _analyze_upload(job_id: str, file_path: str):
    """后台分析上传文件：统计行数 + 抓取预览（不阻塞上传响应）"""
    try:
        t0 = time.time()
        logger.info(f"_analyze_upload start: job={job_id[:8]}")
        total_rows, preview_rows = _fast_preview(file_path)
        update_job(job_id, status="uploaded", total_rows=total_rows,
                   preview_json=json.dumps(preview_rows[:10], ensure_ascii=False))
        logger.info(f"_analyze_upload done: {total_rows} rows in {(time.time()-t0)*1000:.0f}ms")
    except Exception as e:
        update_job(job_id, status="failed", error_msg=f"File analysis failed: {e}")


def _fast_preview(file_path: str) -> tuple[int, list[dict]]:
    """轻量预览：用 openpyxl read_only 只读前 10 行 + 统计总行数。
    201MB Excel 也能秒级完成。"""
    t0 = time.time()
    ext = Path(file_path).suffix.lower()
    try:
        if ext in (".xlsx",):
            import openpyxl
            logger.info(f"_fast_preview opening xlsx: {file_path}")
            t1 = time.time()
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            t2 = time.time()
            logger.info(f"_fast_preview xlsx open: {(t2-t1)*1000:.0f}ms")
            ws = wb.active
            header = None
            preview = []
            total = 0
            for row in ws.iter_rows(values_only=True):
                vals = [str(c) if c is not None else "" for c in row]
                if header is None:
                    header = [chr(65 + i) for i in range(len(vals))]
                    preview.append({header[i]: vals[i] for i in range(len(vals))})
                else:
                    if total < 10:
                        preview.append({header[i]: vals[i] for i in range(min(len(vals), len(header)))})
                total += 1
            wb.close()
            logger.info(f"_fast_preview xlsx done: {total-1} rows in {(time.time()-t0)*1000:.0f}ms total")
            return total - 1, preview[1:] if len(preview) > 1 else []
        elif ext in (".xls",):
            return _fast_preview_xls_or_csv(file_path, engine="xls")
        else:
            return _fast_preview_xls_or_csv(file_path, engine="csv")
    except Exception:
        return 0, []


def _fast_preview_xls_or_csv(file_path: str, engine: str = "csv") -> tuple[int, list[dict]]:
    if engine == "xls":
        import xlrd
        wb = xlrd.open_workbook(file_path)
        ws = wb.sheet_by_index(0)
        header = {chr(65 + ci): str(ws.cell_value(0, ci)).strip() for ci in range(ws.ncols)}
        preview = []
        for ri in range(1, min(ws.nrows, 11)):
            preview.append({chr(65 + ci): str(ws.cell_value(ri, ci)).strip() for ci in range(ws.ncols)})
        return ws.nrows - 1, preview
    else:
        import csv
        # 先数行数（只扫描不解析）
        with open(file_path, encoding="utf-8", errors="replace") as f:
            total = sum(1 for _ in f) - 1  # minus header
        # 再读前 10 行
        with open(file_path, encoding="utf-8", errors="replace") as f:
            reader = csv.reader(f)
            header_row = next(reader, [])
            header = {chr(65 + i): header_row[i] for i in range(len(header_row))}
            preview = []
            for i, row in enumerate(reader):
                if i >= 10:
                    break
                preview.append({chr(65 + j): row[j] for j in range(min(len(row), len(header)))})
        return total, preview


# ═══════════════ Processing ═══════════════

async def _process_job(job_id: str, user_id: str, file_path: str, countries: list[str]):
    """后台处理任务"""
    t0 = time.time()
    logger.info(f"_process_job start: job={job_id[:8]}, file={os.path.basename(file_path)}")
    try:
        update_job(job_id, status="processing")
        user = get_user(user_id)
        job = get_job(job_id)
        file_size_mb = os.path.getsize(file_path) / 1024 / 1024 if os.path.exists(file_path) else 0
        logger.info(f"_process_job file_size={file_size_mb:.1f}MB job.total_rows={job.total_rows}")

        # 配额为 0 时直接拒绝
        if user.quota_remaining <= 0:
            update_job(job_id, status="failed", error_msg="Monthly quota exhausted. Please upgrade to continue.")
            return

        # 调用 pipeline（配额不足时自动截断）
        from .pipeline import run as pipeline_run
        processable = min(user.quota_remaining, job.total_rows)

        def progress_callback(done: int, total: int):
            update_job(job_id, done_rows=done, total_rows=total)

        t_pipeline = time.time()
        result = pipeline_run(excel_path=file_path, countries=countries,
                              max_rows=processable, progress_callback=progress_callback)
        logger.info(f"_process_job pipeline done in {(time.time()-t_pipeline)*1000:.0f}ms")

        ok_count = result.get("ai_full_clean", 0) + result.get("partial_reclean", 0)
        total_skus = result.get("total_sku", 0)
        done = ok_count + result.get("skipped_old", 0) + result.get("price_updated", 0)
        fail_count = total_skus - done

        # 复制 comparison report 到 job 专属文件，供结果页展示
        from shutil import copyfile
        from .config import COMPARISON_REPORT
        job_report = OUTPUT_DIR / f"job_{job_id[:8]}_report.xlsx"
        if COMPARISON_REPORT.exists():
            copyfile(str(COMPARISON_REPORT), str(job_report))

        increment_quota(user_id, ok_count)
        update_job(
            job_id,
            status="completed",
            done_rows=done,
            ok_rows=ok_count,
            fail_rows=max(0, fail_count),
            result_csv=str(job_report),
            truncated=job.total_rows > processable,
        )
    except Exception as e:
        update_job(job_id, status="failed", error_msg=str(e))


@app.post("/api/jobs/{job_id}/process")
async def start_process(
    job_id: str,
    user: User = Depends(current_user),
):
    job = get_job(job_id)
    if not job or job.user_id != user.id:
        raise HTTPException(404, "Job not found")
    if job.status != "uploaded":
        raise HTTPException(400, f"Job cannot be processed in status: {job.status}")

    countries = json.loads(job.country_mask)
    file_id = Path(job.filename).stem
    file_path = None
    for f in UPLOAD_DIR.iterdir():
        if f.name.startswith(job_id[:8]) or f.stem == file_id:
            file_path = str(f)
            break
    if not file_path:
        for f in UPLOAD_DIR.iterdir():
            file_path = str(f)
            break

    if not file_path:
        raise HTTPException(404, "Uploaded file not found")

    import threading
    threading.Thread(target=_process_job, args=(job_id, user.id, file_path, countries), daemon=True).start()
    return {"job_id": job_id, "status": "processing"}


# ═══════════════ Jobs ═══════════════

@app.get("/api/jobs")
async def get_jobs(user: User = Depends(current_user), limit: int = 20):
    jobs = list_jobs(user.id, limit=limit)
    return [
        {
            "id": j.id, "filename": j.filename, "status": j.status,
            "total_rows": j.total_rows, "done_rows": j.done_rows,
            "ok_rows": j.ok_rows, "fail_rows": j.fail_rows,
            "progress_pct": j.progress_pct,
            "created_at": j.created_at,
        }
        for j in jobs
    ]


@app.get("/api/jobs/{job_id}")
async def get_job_detail(job_id: str, user: User = Depends(current_user)):
    job = get_job(job_id)
    if not job or job.user_id != user.id:
        raise HTTPException(404, "Job not found")

    preview_rows = []
    if job.preview_json:
        try:
            preview_rows = json.loads(job.preview_json)
        except (json.JSONDecodeError, TypeError):
            pass

    return {
        "id": job.id, "filename": job.filename, "status": job.status,
        "total_rows": job.total_rows, "preview_rows": preview_rows,
        "done_rows": job.done_rows,
        "ok_rows": job.ok_rows, "fail_rows": job.fail_rows,
        "progress_pct": job.progress_pct,
        "truncated": job.truncated,
        "result_csv": job.result_csv,
        "error_msg": job.error_msg,
        "created_at": job.created_at, "updated_at": job.updated_at,
    }


@app.get("/api/jobs/{job_id}/results")
async def get_job_results(job_id: str, user: User = Depends(current_user)):
    """返回 job 的处理结果预览（前 20 条）"""
    import openpyxl
    job = get_job(job_id)
    if not job or job.user_id != user.id:
        raise HTTPException(404, "Job not found")
    if job.status != "completed":
        return {"rows": [], "message": "Job not finished yet"}

    report_path = Path(job.result_csv) if job.result_csv else None
    if not report_path or not report_path.exists():
        # 回退：尝试从 product_memory 获取最新结果
        from .product_memory import get_recent as get_recent_products
        rows = get_recent_products(limit=20)
        return {"rows": rows, "source": "product_memory"}

    try:
        wb = openpyxl.load_workbook(str(report_path), read_only=True, data_only=True)
        ws = wb.active
        rows = []
        headers = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            vals = [str(c) if c is not None else "" for c in row]
            if i == 0:
                headers = vals
            else:
                rows.append({headers[j]: vals[j] for j in range(min(len(vals), len(headers)))})
                if len(rows) >= 20:
                    break
        wb.close()
        return {"rows": rows, "source": "report", "total": len(rows)}
    except Exception:
        return {"rows": [], "error": "Could not read result file"}


# ═══════════════ Feed Export ═══════════════

@app.get("/api/feeds/{country}")
async def download_feed(country: str, user: User = Depends(current_user)):
    """从 product_memory 动态生成 GMC Feed XML 并返回"""
    from .feed_generator import generate_from_memory

    allowed = {"US", "DE", "FR", "ES", "IT"}
    country = country.upper()
    if country not in allowed:
        raise HTTPException(400, f"Unsupported country: {country}. Allowed: {', '.join(sorted(allowed))}")

    xml = generate_from_memory(country, user_id=user.id)
    return HTMLResponse(content=xml, media_type="application/xml",
                        headers={"Content-Disposition": f"attachment; filename=feed_{country.lower()}.xml"})


@app.get("/api/feeds")
async def list_feeds(user: User = Depends(current_user)):
    """列出可用的 Feed 文件"""
    feeds = []
    for country in ["US", "DE", "FR", "ES", "IT"]:
        feed_path = OUTPUT_DIR / f"feed_{country.lower()}.xml"
        if feed_path.exists():
            stat = feed_path.stat()
            feeds.append({
                "country": country,
                "size_bytes": stat.st_size,
                "updated_at": datetime.fromtimestamp(stat.st_mtime, tz=timezone.utc).isoformat(),
                "download_url": f"/api/feeds/{country}",
            })
    return {"feeds": feeds}


# ═══════════════ PayPal Subscription ═══════════════

PAYPAL_CLIENT_ID = os.getenv("PAYPAL_CLIENT_ID", "")
PAYPAL_PLAN_IDS = {
    "starter": os.getenv("PAYPAL_PLAN_ID_STARTER", ""),
    "growth":  os.getenv("PAYPAL_PLAN_ID_GROWTH", ""),
}

PLAN_QUOTAS = {"starter": 50, "growth": 200}


@app.get("/api/billing/plans")
async def get_paypal_config(user: User = Depends(current_user)):
    """返回 PayPal 配置 + 可选档位"""
    return {
        "client_id": PAYPAL_CLIENT_ID,
        "current_plan": user.plan,
        "plans": {
            plan: {"id": pid, "name": plan.capitalize(), "skus": PLAN_QUOTAS.get(plan, 100)}
            for plan, pid in PAYPAL_PLAN_IDS.items() if pid
        },
    }


@app.post("/api/billing/activate")
async def activate_subscription(
    body: dict,
    user: User = Depends(current_user),
):
    """PayPal 付款成功后激活订阅。
    
    body: {"paypal_subscription_id": "I-XXXX", "paypal_plan_id": "P-XXXX"}
    """
    _require_web_saas()
    sub_id = body.get("paypal_subscription_id", "")
    plan_id = body.get("paypal_plan_id", "")

    if not sub_id:
        raise HTTPException(400, "Missing paypal_subscription_id")

    # 根据 plan_id 找到对应档位
    plan_name = None
    for plan, pid in PAYPAL_PLAN_IDS.items():
        if pid == plan_id:
            plan_name = plan
            break

    if not plan_name:
        raise HTTPException(400, f"Unknown plan_id: {plan_id}")

    from .db import update_user
    quota = PLAN_QUOTAS.get(plan_name, 100)
    update_user(user.id, plan=plan_name, quota_total=quota, stripe_subscription_id=sub_id)

    return {
        "ok": True,
        "plan": plan_name,
        "quota_total": quota,
    }


# ═══════════════ Shopify Integration ═══════════════

from .shopify_client import (
    get_shopify_auth_url, fetch_shopify_products, connect_shopify_store,
)


class ShopifyConnectBody(BaseModel):
    shop_domain: str
    code: str


class ShopifyProcessBody(BaseModel):
    product_ids: list[str]
    countries: list[str] = ["US"]


@app.get("/api/shopify/status")
async def shopify_status(user: User = Depends(current_user)):
    """查询用户是否已连接 Shopify"""
    conn = get_shopify_connection(user.id)
    if conn:
        return {
            "connected": True,
            "shop_domain": conn.shop_domain,
            "shop_name": conn.shop_name,
            "connected_at": conn.created_at,
        }
    return {"connected": False}


@app.get("/api/shopify/auth-url")
async def shopify_auth_url(shop: str = ""):
    """获取 Shopify OAuth 授权 URL"""
    if not shop:
        raise HTTPException(400, "Provide a Shopify shop domain, e.g. mystore")
    url = get_shopify_auth_url(shop)
    return {"url": url}


@app.get("/api/shopify/callback")
async def shopify_callback(shop: str = "", code: str = ""):
    """Shopify OAuth 回调 — 换 token 并存储"""
    if not shop or not code:
        raise HTTPException(400, "Missing shop or code parameter")

    # 这里需要从 cookie 或 session 获取用户
    # 简化处理：要求前端在 callback 页面手动调用 /api/shopify/connect
    return {"shop": shop, "code": code, "message": "Complete connection from the frontend"}


@app.post("/api/shopify/connect")
async def shopify_connect(body: ShopifyConnectBody, user: User = Depends(current_user)):
    """连接 Shopify 店铺（前端拿到 code 后调用）"""
    result = await connect_shopify_store(user.id, body.shop_domain, body.code)
    if not result:
        raise HTTPException(400, "Shopify authorization failed — check shop domain and code")

    # Also mirror into stores (App single source of truth)
    from . import store_db
    from .db import get_shopify_connection
    conn = get_shopify_connection(user.id)
    if conn:
        existing = store_db.get_store_by_domain(conn.shop_domain)
        if existing:
            store_db.update_store(
                existing.id,
                access_token=conn.access_token,
                shop_name=conn.shop_name,
                status="active",
            )
        else:
            store_db.create_store(
                user_id=user.id,
                shopify_domain=conn.shop_domain,
                shop_name=conn.shop_name,
                access_token=conn.access_token,
            )
    return result


@app.post("/api/shopify/disconnect")
async def shopify_disconnect(user: User = Depends(current_user)):
    """断开 Shopify 连接"""
    deleted = delete_shopify_connection(user.id)
    return {"ok": deleted}


@app.get("/api/shopify/products")
async def shopify_products(
    page_info: str = "",
    limit: int = 50,
    user: User = Depends(current_user),
):
    """拉取 Shopify 产品列表"""
    conn = get_shopify_connection(user.id)
    if not conn:
        raise HTTPException(400, "Connect a Shopify store first")

    result = await fetch_shopify_products(
        shop_domain=conn.shop_domain,
        access_token=conn.access_token,
        limit=limit,
        page_info=page_info or None,
    )
    return result


@app.post("/api/shopify/process")
async def shopify_process(body: ShopifyProcessBody, user: User = Depends(current_user)):
    """选择 Shopify 产品 + 国家 → 启动 AI pipeline"""
    conn = get_shopify_connection(user.id)
    if not conn:
        raise HTTPException(400, "Connect a Shopify store first")

    if not body.product_ids:
        raise HTTPException(400, "Select at least one product")

    # 拉取选中产品的完整数据
    all_products = await fetch_shopify_products(
        shop_domain=conn.shop_domain,
        access_token=conn.access_token,
        limit=250,
    )

    # 过滤出用户选中的产品
    selected = [p for p in all_products["products"] if p["shopify_id"] in body.product_ids]
    if not selected:
        raise HTTPException(404, "Selected products not found")

    # 创建 job
    job = create_job(
        user.id,
        f"shopify_{conn.shop_name}_{len(selected)}_products",
        json.dumps(body.countries),
        f"shopify_{conn.shop_domain}",
    )

    # 后台线程处理
    import threading
    threading.Thread(
        target=_process_shopify_job,
        args=(job.id, user.id, selected, body.countries),
        daemon=True,
    ).start()

    return {"job_id": job.id, "status": "processing", "product_count": len(selected)}


def _process_shopify_job(job_id: str, user_id: str, products: list, countries: list):
    """后台处理 Shopify 产品 job"""
    from .pipeline import run as pipeline_run
    try:
        logger.info(f"Shopify job start: {job_id[:8]} ({len(products)} products, {countries})")
        update_job(job_id, status="processing")

        def progress_cb(done, total):
            update_job(job_id, done_rows=done, total_rows=total)

        result = pipeline_run(
            products_data=products,
            countries=countries,
            progress_callback=progress_cb,
        )

        update_job(job_id, status="completed", total_rows=result.get("total_sku", 0),
                   ok_rows=result.get("ai_full_clean", 0) + result.get("partial_reclean", 0))
        logger.info(f"Shopify job done: {job_id[:8]} — {result.get('total_sku', 0)} SKU")
    except Exception as e:
        logger.error(f"Shopify job failed: {job_id[:8]} — {e}")
        update_job(job_id, status="failed", error_msg=str(e))


# ═══════════════ Shopify Feed API (Plugin) ═══════════════

class ShopifyFeedRequest(BaseModel):
    product_ids: list[str]
    countries: list[str] = ["US"]
    shop_domain: str


@app.post("/api/shopify/feed")
async def shopify_generate_feed(body: ShopifyFeedRequest):
    """LEGACY — unauthenticated generate disabled. Use /api/app/generate with session."""
    raise HTTPException(
        410,
        "This endpoint is retired. Use authenticated /api/app/generate with a Shopify session token.",
    )


def _xml_to_csv(xml_path: Path, csv_path: Path):
    """将 Feed XML 转为 TSV（Google 标准 Tab-delimited 格式）"""
    import re
    content = xml_path.read_text(encoding="utf-8")
    items = re.findall(r'<item>(.*?)</item>', content, re.DOTALL)
    if not items:
        return

    # 提取所有 g: 命名空间字段
    fields = set()
    rows = []
    for item in items:
        row = {}
        # 普通字段
        for tag in ["id", "title", "description", "link", "image_link",
                    "price", "availability", "condition", "brand",
                    "gtin", "mpn", "identifier_exists"]:
            m = re.search(f'<g:{tag}>(.*?)</g:{tag}>', item)
            if m:
                row[tag] = m.group(1)
                fields.add(tag)
        # 可能重复的字段
        for tag in ["additional_image_link", "product_highlight"]:
            vals = re.findall(f'<g:{tag}>(.*?)</g:{tag}>', item)
            if vals:
                row[tag] = ",".join(vals)
                fields.add(tag)
        # 变体字段
        for tag in ["item_group_id", "color", "size", "size_system",
                    "gender", "age_group", "product_type"]:
            m = re.search(f'<g:{tag}>(.*?)</g:{tag}>', item)
            if m:
                row[tag] = m.group(1)
                fields.add(tag)
        # custom labels
        for i in range(5):
            tag = f"custom_label_{i}"
            m = re.search(f'<g:{tag}>(.*?)</g:{tag}>', item)
            if m:
                row[tag] = m.group(1)
                fields.add(tag)
        rows.append(row)

    # 写 TSV
    import csv
    ordered_fields = sorted(fields)
    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=ordered_fields, delimiter="\t",
                                extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow(row)

    logger.info(f"CSV generated: {csv_path} ({len(rows)} rows)")


@app.get("/api/shopify/feed/status")
async def shopify_feed_status(shop_domain: str = ""):
    """查询店铺 Feed 状态（URL、商品数、更新时间）"""
    from . import store_db

    if not shop_domain:
        raise HTTPException(400, "shop_domain is required")

    shop_domain = shop_domain.replace(".myshopify.com", "").strip() + ".myshopify.com"
    store = store_db.get_store_by_domain(shop_domain)
    if not store:
        return {"feeds": []}

    feed_files = store_db.list_store_feeds(store.id)
    feeds = []
    for ff in feed_files:
        feeds.append({
            "country": ff.country,
            "url": ff.feed_url,
            "csv_url": ff.feed_url.replace(".xml", ".csv"),
            "item_count": ff.item_count,
            "updated_at": ff.generated_at,
        })

    return {"feeds": feeds, "store_id": store.id}


# 静态文件服务：Feed XML/CSV 下载（带 Cache-Control 和 CORS 头）
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse
from .config import FEEDS_DIR

FEEDS_DIR.mkdir(parents=True, exist_ok=True)


def _feed_file_response(file_path: Path, filename: str):
    if not file_path.exists() or not file_path.is_file():
        raise HTTPException(404, f"Feed file not found: {filename}")
    if filename.endswith(".xml"):
        media_type = "application/xml"
    elif filename.endswith(".csv"):
        media_type = "text/csv"
    else:
        media_type = "application/octet-stream"
    response = FileResponse(
        path=str(file_path),
        media_type=media_type,
        filename=filename,
    )
    response.headers["Cache-Control"] = "public, max-age=3600, must-revalidate"
    response.headers["Access-Control-Allow-Origin"] = "*"
    response.headers["X-Content-Type-Options"] = "nosniff"
    return response


@app.get("/feeds/{store_id}/{platform}/{filename}")
async def serve_platform_feed_file(store_id: str, platform: str, filename: str):
    """Public durable feed: /feeds/{store}/{platform}/{lang}.xml|csv"""
    if ".." in store_id or ".." in platform or ".." in filename:
        raise HTTPException(400, "Invalid path")
    return _feed_file_response(FEEDS_DIR / store_id / platform / filename, filename)


@app.get("/feeds/{store_id}/{filename}")
async def serve_feed_file(store_id: str, filename: str):
    """Legacy flat path + GMC polling fallback."""
    if ".." in store_id or ".." in filename:
        raise HTTPException(400, "Invalid path")
    return _feed_file_response(FEEDS_DIR / store_id / filename, filename)


# ── Shopify catalog / uninstall / GDPR webhooks ──

@app.post("/api/webhooks/shopify/products_update")
@app.post("/api/webhooks/shopify/products_update/")
async def webhook_products_update(request: Request):
    from .shopify_webhooks import verify_shopify_hmac, handle_products_update
    raw = await request.body()
    if not verify_shopify_hmac(raw, request.headers.get("X-Shopify-Hmac-Sha256", "")):
        if config.SHOPIFY_CLIENT_SECRET and os.getenv("ADFEED_WEBHOOK_SKIP_HMAC", "").lower() not in ("1", "true", "yes"):
            raise HTTPException(401, "Invalid webhook HMAC")
    payload = json.loads(raw.decode("utf-8") or "{}")
    shop = request.headers.get("X-Shopify-Shop-Domain", "")
    return handle_products_update(shop, payload)


@app.post("/api/webhooks/shopify/products_delete")
@app.post("/api/webhooks/shopify/products_delete/")
async def webhook_products_delete(request: Request):
    from .shopify_webhooks import verify_shopify_hmac, handle_products_delete
    raw = await request.body()
    if not verify_shopify_hmac(raw, request.headers.get("X-Shopify-Hmac-Sha256", "")):
        if config.SHOPIFY_CLIENT_SECRET and os.getenv("ADFEED_WEBHOOK_SKIP_HMAC", "").lower() not in ("1", "true", "yes"):
            raise HTTPException(401, "Invalid webhook HMAC")
    payload = json.loads(raw.decode("utf-8") or "{}")
    shop = request.headers.get("X-Shopify-Shop-Domain", "")
    return handle_products_delete(shop, payload)


@app.post("/api/webhooks/shopify/app_uninstalled")
@app.post("/api/webhooks/shopify/app_uninstalled/")
async def webhook_app_uninstalled(request: Request):
    from .shopify_webhooks import verify_shopify_hmac, handle_app_uninstalled
    raw = await request.body()
    if not verify_shopify_hmac(raw, request.headers.get("X-Shopify-Hmac-Sha256", "")):
        if config.SHOPIFY_CLIENT_SECRET and os.getenv("ADFEED_WEBHOOK_SKIP_HMAC", "").lower() not in ("1", "true", "yes"):
            raise HTTPException(401, "Invalid webhook HMAC")
    shop = request.headers.get("X-Shopify-Shop-Domain", "")
    return handle_app_uninstalled(shop)


async def _webhook_gdpr(request: Request):
    from .shopify_webhooks import verify_shopify_hmac, handle_compliance_webhook
    raw = await request.body()
    hmac_header = request.headers.get("X-Shopify-Hmac-Sha256", "")
    if not verify_shopify_hmac(raw, hmac_header):
        if config.SHOPIFY_CLIENT_SECRET and os.getenv("ADFEED_WEBHOOK_SKIP_HMAC", "").lower() not in ("1", "true", "yes"):
            return Response(status_code=401)
    topic = request.headers.get("X-Shopify-Topic", "gdpr")
    try:
        payload = json.loads(raw.decode("utf-8") or "{}")
    except json.JSONDecodeError:
        payload = {}
    shop = request.headers.get("X-Shopify-Shop-Domain", "") or payload.get("shop_domain", "")
    return handle_compliance_webhook(topic, payload, shop)


async def _webhook_gdpr_entry(request: Request):
    return await _webhook_gdpr(request)


# Partner automated checks may hit the unified URI or legacy per-topic paths (with/without trailing slash).
_GDPR_WEBHOOK_PATHS = [
    "/api/webhooks/shopify/compliance",
    "/api/webhooks/shopify/customers/data_request",
    "/api/webhooks/shopify/customers/redact",
    "/api/webhooks/shopify/shop/redact",
    "/api/webhooks/shopify/customers_redact",
    "/api/webhooks/shopify/shop_redact",
    # Stale checker URLs (no /api prefix) seen in production nginx logs.
    "/webhooks/compliance",
    "/webhooks/customers/data_request",
    "/webhooks/customers/data/request",
    "/webhooks/customers/redact",
    "/webhooks/shop/redact",
    "/customers/data_request",
    "/customers/data/request",
    "/customers/redact",
    "/shop/redact",
]
for _gdpr_path in _GDPR_WEBHOOK_PATHS:
    app.add_api_route(_gdpr_path, _webhook_gdpr_entry, methods=["POST"])
    app.add_api_route(_gdpr_path + "/", _webhook_gdpr_entry, methods=["POST"])


# 保留 StaticFiles 作为回退（处理其他路径）
app.mount("/feeds-static", StaticFiles(directory=str(FEEDS_DIR)), name="feeds")
